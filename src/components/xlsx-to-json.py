#!/usr/bin/python3

"""

extracts key pieces from Excel spreadsheets into JSON

Emery Berger
Microsoft Research / University of Massachusetts Amherst
http://www.emeryberger.com

"""

import json
import time
import random
import openpyxl
import io
import os
import sys
import argparse
import warnings
import pprint
import hashlib
import shutil
import tempfile

# From Stack Overflow:
# https://stackoverflow.com/questions/23562366/how-do-i-get-value-present-in-a-merged-cell

from openpyxl.utils import get_column_letter

def getValueWithMergeLookup(sheet, cell):
    idx = cell.coordinate
    for range_ in sheet.merged_cells.ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(range_))
        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range
                return sheet.cell(merged_cells[0][0]).value


with warnings.catch_warnings():
    warnings.simplefilter("ignore")

# process a single worksheet
#   datatype = "f" --> functions
#   datatype = "n" --> numeric values

# From Stack Overflow:
# https://stackoverflow.com/questions/53852149/how-do-i-find-the-last-nonblank-cell-in-openpyxl


def find_edges(sheet):
    row = sheet.max_row
    while row > 0:
        cells = sheet[row]
        if all([cell.value is None for cell in cells]):
            row -= 1
        else:
            break
    if row == 0:
        return 0, 0

    column = sheet.max_column
    while column > 0:
        cells = next(sheet.iter_cols(
            min_col=column, max_col=column, max_row=row))
        if all([cell.value is None for cell in cells]):
            column -= 1
        else:
            break
    return row, column


def find_rows(sheet, col):
    rows = sheet.iter_rows(max_col=col)
    row = 1
    ri = 1
    while True:
        try:
            cells = next(rows)
            if not all([type(cell).__name__ == 'MergedCell' or cell.value is None for cell in cells]):
                row = ri
            ri += 1
        except StopIteration:
            break
    return row


def find_columns(sheet, row):
    columns = sheet.iter_cols(max_row=row)
    column = 1
    ci = 1
    while True:
        try:
            cells = next(columns)
            if not all([type(cell).__name__ == 'MergedCell' or cell.value is None for cell in cells]):
                column = ci
            ci += 1
        except StopIteration:
            break
    return column


#            if type(cell).__name__ != 'MergedCell':
#                if cell.value is not None:
#                    if cell.data_type == datatype:
#                        v = str(cell.value)

def process_sheet(worksheet, datatype="f"):
    processed = []
    theRows = worksheet.rows
    nRow = 0
    for row in theRows:
        nRow += 1
        r = []
        nCol = 0
        for cell in row:
            nCol += 1
            v = ""
            if type(cell).__name__ != 'MergedCell':
                if cell.value is not None:
                    if cell.data_type == datatype:
                        v = str(cell.value)
            r.append(v)
        processed.append(r)
    return processed

# Get styles


properties = {'format': {
    'fill': {
        'color': True
    },
    'border': {
        'color': True,
        'style': True,
        'weight': True
    },
    'font': {
        'color': True,
        'style': True,
        'weight': True,
        'bold': True,
        'italic': True,
        'name': True
    }
}


}


def process_sheet_styles(worksheet):
    processed = []
    pp = pprint.PrettyPrinter(indent=8, depth=8)
    for row in worksheet.rows:
        r = []
        for cell in row:
            v = ""
            if type(cell).__name__ != 'MergedCell':
                # Instead of storing the string here, store a compact hash.
                styles = str(vars(cell.font)) + \
                    str(vars(cell.fill)) + str(vars(cell.border))
                v = hashlib.sha224(styles.encode('utf-8')).hexdigest()
                # Only grab first 10 characters; should be enough probabilistically.
                v = v[0:9]
                # v = str([vars(cell.font), vars(cell.fill), vars(cell.border)])
            r.append(v)
        processed.append(r)
    return processed


# Process every sheet in a workbook
def process_workbook(dirname, fname):
    qualified_origname = os.path.join(dirname, fname)
    with open(qualified_origname, 'rb') as f:
        in_mem_file = io.BytesIO(f.read())
    readonly = False
    try:
        workbook = openpyxl.load_workbook(in_mem_file, read_only=readonly)
    except:
        # OK, we couldn't open it R/W. Try to process R/O.
        readonly = True
        workbook = openpyxl.load_workbook(in_mem_file, read_only=readonly)
    output = {
        "workbookName": fname,
        "worksheets": []
    }
    trimmed = False
    for worksheet in workbook.worksheets:
        
#        if not readonly:
#            worksheet.reset_dimensions()
            
        sheetname = worksheet.title
        # Now get the sheet range info.
        numRows = worksheet.max_row
        numCols = worksheet.max_column

        if numRows > 100 or numCols > 100:
            origRows = numRows
            origCols = numCols
            trimmed = True
            # Big sheet.
            # Worth the effort to see if we can trim it.
            if numRows < numCols:
                numCols = find_columns(worksheet, numRows)
                numRows = find_rows(worksheet, numCols)
            else:
                numRows = find_rows(worksheet, numCols)
                numCols = find_columns(worksheet, numRows)
            # Now trim the sheet.
            if origCols != numCols:
                worksheet.delete_cols(numCols + 1, origCols - numCols)
            if origRows != numRows:
                worksheet.delete_rows(numRows + 1, origRows - numRows)
        
#        theRows = 
#        print(tuple(worksheet.rows))
        
        # Extract formulas and numeric values.
#        print("----VALUES----")
        values = process_sheet(worksheet, "n")
#        print("----DONE WITH VALUES-----")
#        print("--- FORMULAS ----")
        formulas = process_sheet(worksheet, "f")
#        print("--- DONE WITH FORMULAS---")
        
        # Get the styles of each cell.
        styles = process_sheet_styles(worksheet)

        lastColName = openpyxl.utils.cell.get_column_letter(numCols)
        sheetRange = sheetname + "!A1:" + lastColName + str(numRows)

        output["worksheets"].append({
            "sheetName": sheetname,
            "usedRangeAddress": sheetRange,
            "formulas": formulas,
            "values": values,
            "styles": styles
        })

    # Save if trimmed
#    if trimmed:
#        workbook.save(fname)

    return output


# Process command line for file or directory.

parser = argparse.ArgumentParser('xlsx-to-json.py')
parser.add_argument('--input', help='Process an input .xlsx file.')
# TODO: specify file output
parser.add_argument('--directory', help='Process all .xlsx files in a directory.')

args = parser.parse_args()

if args.input is None and args.directory is None:
    parser.print_help()
    exit(-1)

if args.directory is None:
    output = process_workbook('.', args.input)
    s = json.dumps(output)
    print(s)
else:
    fnames = os.listdir(args.directory)
    for fname in fnames:
        # Skip subdirectories
        if os.path.isdir(os.path.join(args.directory, fname)):
            continue
        if fname.endswith("xlsx") and not fname.startswith("~$"):
            print("processing " + fname)
            try:
                output = process_workbook(args.directory, fname)
                s = json.dumps(output)
                output_fname = fname.replace(".xlsx", ".json")
                with open(os.path.join(args.directory, output_fname), "w") as f:
                    f.write(s)
            except:
                pass

        
